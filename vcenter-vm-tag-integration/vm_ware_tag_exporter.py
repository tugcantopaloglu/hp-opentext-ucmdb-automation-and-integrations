#!/usr/bin/env python3
"""
VMware Tag Exporter
Exports VMware VM names and tags to an Excel file for use with VMware-SMAX Bridge

Usage:
    python vmware_tag_exporter.py export --config config.json --output vms.xlsx
    python vmware_tag_exporter.py export --config config.json --output vms.xlsx --category sorumlu
    python vmware_tag_exporter.py show --config config.json
    python vmware_tag_exporter.py template --output manual_vms.xlsx
"""

import argparse
import json
import logging
import sys
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional

import requests
import urllib3

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

__version__ = "1.0.0"

logger = logging.getLogger(__name__)


# =============================================================================
# CONFIGURATION
# =============================================================================


@dataclass
class VMwareConfig:
    """VMware vCenter connection configuration"""
    host: str
    username: str
    password: str
    port: int = 443
    disable_ssl_verification: bool = False
    name: str = ""

    def __post_init__(self):
        if not self.name:
            self.name = self.host


@dataclass
class ExportConfig:
    """Export configuration"""
    owner_tag_category: str = "sorumlu"
    include_all_tags: bool = False
    include_power_state: bool = True


# =============================================================================
# VMWARE CLIENT (REST API)
# =============================================================================


class VMwareClient:
    """Client for VMware vCenter REST API"""

    def __init__(self, config: VMwareConfig):
        self.config = config
        self.session = None
        self.base_url = f"https://{config.host}/api"
        self._session_id = None

    def connect(self) -> None:
        """Connect to vCenter"""
        from requests.auth import HTTPBasicAuth

        self.session = requests.Session()
        self.session.verify = not self.config.disable_ssl_verification

        auth_url = f"{self.base_url}/session"
        response = self.session.post(
            auth_url, auth=HTTPBasicAuth(self.config.username, self.config.password)
        )
        response.raise_for_status()

        self._session_id = response.json()
        self.session.headers.update({"vmware-api-session-id": self._session_id})

        logger.info(f"Connected to vCenter: {self.config.host}")

    def disconnect(self) -> None:
        """Disconnect from vCenter"""
        if self.session and self._session_id:
            try:
                self.session.delete(f"{self.base_url}/session")
            except Exception:
                pass
            logger.info(f"Disconnected from vCenter: {self.config.name}")

    def get_all_vms(self) -> List[Dict[str, Any]]:
        """Get all VMs"""
        response = self.session.get(f"{self.base_url}/vcenter/vm")
        response.raise_for_status()
        vms = response.json()
        logger.info(f"Retrieved {len(vms)} VMs from {self.config.name}")
        return vms

    def get_tag_categories(self) -> Dict[str, str]:
        """Get all tag categories: {category_id: category_name}"""
        response = self.session.get(f"{self.base_url}/cis/tagging/category")
        response.raise_for_status()

        categories = {}
        for category_id in response.json():
            cat_response = self.session.get(
                f"{self.base_url}/cis/tagging/category/{category_id}"
            )
            if cat_response.ok:
                cat_data = cat_response.json()
                categories[category_id] = cat_data.get("name", category_id)

        logger.info(f"Retrieved {len(categories)} tag categories")
        return categories

    def get_tags(self) -> Dict[str, Dict[str, Any]]:
        """Get all tags: {tag_id: {name, category_id}}"""
        response = self.session.get(f"{self.base_url}/cis/tagging/tag")
        response.raise_for_status()

        tags = {}
        for tag_id in response.json():
            tag_response = self.session.get(f"{self.base_url}/cis/tagging/tag/{tag_id}")
            if tag_response.ok:
                tag_data = tag_response.json()
                tags[tag_id] = {
                    "name": tag_data.get("name"),
                    "category_id": tag_data.get("category_id"),
                }

        logger.info(f"Retrieved {len(tags)} tags")
        return tags

    def get_all_vm_tags(self) -> Dict[str, List[str]]:
        """Get tags for all VMs: {vm_id: [tag_ids]}"""
        response = self.session.get(f"{self.base_url}/cis/tagging/tag")
        response.raise_for_status()

        vm_tags = {}
        tag_ids = response.json()

        for tag_id in tag_ids:
            attached_response = self.session.post(
                f"{self.base_url}/cis/tagging/tag-association/{tag_id}?action=list-attached-objects"
            )

            if attached_response.ok:
                for obj in attached_response.json():
                    if obj.get("type") == "VirtualMachine":
                        vm_id = obj.get("id")
                        if vm_id not in vm_tags:
                            vm_tags[vm_id] = []
                        vm_tags[vm_id].append(tag_id)

        logger.info(f"Retrieved tags for {len(vm_tags)} VMs")
        return vm_tags

    def __enter__(self):
        self.connect()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.disconnect()


# =============================================================================
# EXPORTER
# =============================================================================


class VMwareTagExporter:
    """Exports VMware VM tags to Excel"""

    def __init__(self, configs: List[VMwareConfig], export_config: ExportConfig):
        self.configs = configs
        self.export_config = export_config

    def collect_data(self) -> List[Dict[str, Any]]:
        """Collect VM and tag data from all vCenters"""
        all_data = []

        for config in self.configs:
            try:
                with VMwareClient(config) as client:
                    # Get metadata
                    categories = client.get_tag_categories()
                    tags = client.get_tags()
                    vms = client.get_all_vms()
                    vm_tags = client.get_all_vm_tags()

                    # Build category name lookup
                    category_names = {cat_id: name for cat_id, name in categories.items()}

                    # Build tag lookup with category names
                    tag_lookup = {}
                    for tag_id, tag_data in tags.items():
                        cat_id = tag_data["category_id"]
                        cat_name = category_names.get(cat_id, "Unknown")
                        tag_lookup[tag_id] = {
                            "name": tag_data["name"],
                            "category": cat_name,
                        }

                    # Process each VM
                    for vm in vms:
                        vm_id = vm.get("vm", "")
                        vm_name = vm.get("name", "")
                        power_state = vm.get("power_state", "UNKNOWN")

                        # Get tags for this VM
                        vm_tag_ids = vm_tags.get(vm_id, [])

                        # Organize tags by category
                        tags_by_category = {}
                        for tag_id in vm_tag_ids:
                            if tag_id in tag_lookup:
                                tag_info = tag_lookup[tag_id]
                                cat_name = tag_info["category"]
                                tag_name = tag_info["name"]

                                if cat_name not in tags_by_category:
                                    tags_by_category[cat_name] = []
                                tags_by_category[cat_name].append(tag_name)

                        # Get owner tag
                        owner_tags = tags_by_category.get(self.export_config.owner_tag_category, [])
                        owner_tag = owner_tags[0] if owner_tags else ""

                        row = {
                            "vm_name": vm_name,
                            "owner_tag": owner_tag,
                            "vcenter": config.name,
                            "power_state": power_state,
                            "all_tags": tags_by_category,
                        }

                        all_data.append(row)

                    logger.info(f"Collected {len(vms)} VMs from {config.name}")

            except Exception as e:
                logger.error(f"Failed to collect from {config.name}: {e}")

        return all_data

    def export_to_excel(self, data: List[Dict[str, Any]], output_path: str) -> None:
        """Export data to Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            sys.exit(1)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VM Tags"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers - These column names are used by vmware_smax_bridge.py
        headers = ["VM Name", "Owner", "Source", "Power State"]
        if self.export_config.include_all_tags:
            headers.append("All Tags")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data rows
        for row_idx, row_data in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=row_data["vm_name"]).border = border
            ws.cell(row=row_idx, column=2, value=row_data["owner_tag"]).border = border
            ws.cell(row=row_idx, column=3, value=row_data["vcenter"]).border = border
            ws.cell(row=row_idx, column=4, value=row_data["power_state"]).border = border

            if self.export_config.include_all_tags:
                tags_str = "; ".join(
                    f"{cat}: {', '.join(tags)}"
                    for cat, tags in row_data["all_tags"].items()
                )
                ws.cell(row=row_idx, column=5, value=tags_str).border = border

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 40  # VM Name
        ws.column_dimensions['B'].width = 15  # Owner
        ws.column_dimensions['C'].width = 25  # Source
        ws.column_dimensions['D'].width = 15  # Power State
        if self.export_config.include_all_tags:
            ws.column_dimensions['E'].width = 50

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add metadata sheet
        ws_meta = wb.create_sheet(title="Metadata")
        ws_meta["A1"] = "Export Date"
        ws_meta["B1"] = datetime.now().isoformat()
        ws_meta["A2"] = "Total VMs"
        ws_meta["B2"] = len(data)
        ws_meta["A3"] = "Owner Tag Category"
        ws_meta["B3"] = self.export_config.owner_tag_category
        ws_meta["A4"] = "vCenters"
        ws_meta["B4"] = ", ".join(c.name for c in self.configs)

        # Count stats
        with_owner = sum(1 for d in data if d["owner_tag"])
        without_owner = len(data) - with_owner
        ws_meta["A5"] = "VMs with owner tag"
        ws_meta["B5"] = with_owner
        ws_meta["A6"] = "VMs without owner tag"
        ws_meta["B6"] = without_owner

        wb.save(output_path)
        logger.info(f"Exported {len(data)} VMs to {output_path}")

    def show_summary(self, data: List[Dict[str, Any]]) -> None:
        """Show summary of collected data"""
        print("\n" + "=" * 60)
        print("VMware Tag Export Summary")
        print("=" * 60)

        total = len(data)
        with_owner = sum(1 for d in data if d["owner_tag"])
        without_owner = total - with_owner

        print(f"Total VMs: {total}")
        print(f"VMs with '{self.export_config.owner_tag_category}' tag: {with_owner}")
        print(f"VMs without '{self.export_config.owner_tag_category}' tag: {without_owner}")

        # By vCenter
        by_vcenter = {}
        for d in data:
            vc = d["vcenter"]
            if vc not in by_vcenter:
                by_vcenter[vc] = {"total": 0, "with_owner": 0}
            by_vcenter[vc]["total"] += 1
            if d["owner_tag"]:
                by_vcenter[vc]["with_owner"] += 1

        print("\nBy vCenter:")
        print("-" * 40)
        for vc, stats in by_vcenter.items():
            pct = (stats["with_owner"] / stats["total"] * 100) if stats["total"] > 0 else 0
            print(f"  {vc}: {stats['total']} VMs ({stats['with_owner']} tagged, {pct:.1f}%)")

        # Sample VMs without owner
        untagged = [d for d in data if not d["owner_tag"]][:10]
        if untagged:
            print(f"\nSample VMs without '{self.export_config.owner_tag_category}' tag:")
            print("-" * 40)
            for d in untagged:
                print(f"  [{d['vcenter']}] {d['vm_name']}")
            if without_owner > 10:
                print(f"  ... and {without_owner - 10} more")

        print("=" * 60 + "\n")


def create_template(output_path: str) -> None:
    """Create an empty Excel template for manual entry"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VM Tags"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers - must match what vmware_smax_bridge.py expects
    headers = ["VM Name", "Owner", "Source", "Power State"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Example rows
    examples = [
        ("SERVER01", "z12345", "manual", "POWERED_ON"),
        ("SERVER02", "z67890", "manual", "POWERED_ON"),
        ("DATABASE01", "z11111", "manual", "POWERED_ON"),
    ]

    for row_idx, (vm_name, owner, source, power) in enumerate(examples, 2):
        ws.cell(row=row_idx, column=1, value=vm_name).border = border
        ws.cell(row=row_idx, column=2, value=owner).border = border
        ws.cell(row=row_idx, column=3, value=source).border = border
        ws.cell(row=row_idx, column=4, value=power).border = border

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15

    # Freeze header
    ws.freeze_panes = "A2"

    # Instructions sheet
    inst = wb.create_sheet(title="Instructions")
    instructions = [
        "VMware-SMAX Bridge Excel Input Template",
        "",
        "This Excel file can be used as an additional data source for vmware_smax_bridge.py",
        "",
        "Columns:",
        "  - VM Name: The name of the VM (must match SMAX CI DisplayLabel or HostName)",
        "  - Owner: The owner's UPN (e.g., z12345) - must match SMAX Person Upn",
        "  - Source: Identifier for the data source (e.g., 'manual', 'excel')",
        "  - Power State: POWERED_ON or POWERED_OFF (optional, for reference)",
        "",
        "Usage with vmware_smax_bridge.py:",
        "  python vmware_smax_bridge.py sync --config config.json --excel vms.xlsx",
        "",
        "Notes:",
        "  - Excel entries are merged with VMware vCenter data",
        "  - If a VM exists in both Excel and vCenter, Excel takes precedence",
        "  - Only rows with both VM Name and Owner will be processed",
    ]

    for row_idx, line in enumerate(instructions, 1):
        cell = inst.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)

    inst.column_dimensions['A'].width = 80

    wb.save(output_path)
    print(f"✓ Template created: {output_path}")


# =============================================================================
# EXCEL READER (for use by vmware_smax_bridge.py)
# =============================================================================


def read_vm_owners_from_excel(excel_path: str) -> List[Dict[str, str]]:
    """
    Read VM-owner mappings from Excel file.
    Returns list of dicts: [{"vm_name": "...", "owner": "...", "source": "..."}]
    
    This function is imported and used by vmware_smax_bridge.py
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed")
        return []

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        # Find column indices from header row
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                header_lower = str(cell.value).lower().strip()
                if "vm" in header_lower and "name" in header_lower:
                    headers["vm_name"] = col_idx
                elif header_lower == "owner" or "owner" in header_lower:
                    headers["owner"] = col_idx
                elif header_lower == "source" or "vcenter" in header_lower:
                    headers["source"] = col_idx

        if "vm_name" not in headers or "owner" not in headers:
            logger.warning(f"Excel file missing required columns. Found: {list(headers.keys())}")
            return []

        # Read data rows
        mappings = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            vm_name = row[headers["vm_name"] - 1] if headers["vm_name"] - 1 < len(row) else None
            owner = row[headers["owner"] - 1] if headers["owner"] - 1 < len(row) else None
            source = "excel"
            if "source" in headers and headers["source"] - 1 < len(row):
                source = row[headers["source"] - 1] or "excel"

            # Only include rows with both vm_name and owner
            if vm_name and owner:
                mappings.append({
                    "vm_name": str(vm_name).strip(),
                    "owner": str(owner).strip(),
                    "source": str(source).strip(),
                })

        wb.close()
        logger.info(f"Read {len(mappings)} VM-owner mappings from {excel_path}")
        return mappings

    except Exception as e:
        logger.error(f"Failed to read Excel file {excel_path}: {e}")
        return []


# =============================================================================
# CLI
# =============================================================================


def setup_logging(verbose: bool = False) -> None:
    """Configure logging"""
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )


def load_config(config_path: str) -> List[VMwareConfig]:
    """Load VMware configuration from file"""
    with open(config_path, "r") as f:
        data = json.load(f)

    vmware_data = data.get("vmware", {})

    if isinstance(vmware_data, list):
        return [
            VMwareConfig(
                host=vc.get("host", ""),
                username=vc.get("username", ""),
                password=vc.get("password", ""),
                port=vc.get("port", 443),
                disable_ssl_verification=vc.get("disable_ssl_verification", False),
                name=vc.get("name", ""),
            )
            for vc in vmware_data
        ]
    else:
        return [
            VMwareConfig(
                host=vmware_data.get("host", ""),
                username=vmware_data.get("username", ""),
                password=vmware_data.get("password", ""),
                port=vmware_data.get("port", 443),
                disable_ssl_verification=vmware_data.get("disable_ssl_verification", False),
                name=vmware_data.get("name", ""),
            )
        ]


def main():
    parser = argparse.ArgumentParser(
        description="VMware Tag Exporter - Export VM tags to Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export VMs from vCenters to Excel
  python vmware_tag_exporter.py export --config config.json --output vms.xlsx

  # Export with all tags
  python vmware_tag_exporter.py export --config config.json --output vms.xlsx --all-tags

  # Show summary without exporting
  python vmware_tag_exporter.py show --config config.json

  # Create empty template for manual entry
  python vmware_tag_exporter.py template --output manual_vms.xlsx
        """,
    )

    subparsers = parser.add_subparsers(dest="command", help="Available commands")

    # Export command
    export_parser = subparsers.add_parser("export", help="Export VM tags to Excel")
    export_parser.add_argument("--config", "-c", required=True, help="Path to config JSON")
    export_parser.add_argument("--output", "-o", required=True, help="Output Excel file path")
    export_parser.add_argument("--category", default="sorumlu", help="Owner tag category (default: sorumlu)")
    export_parser.add_argument("--all-tags", action="store_true", help="Include all tags")
    export_parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")

    # Show command
    show_parser = subparsers.add_parser("show", help="Show summary without exporting")
    show_parser.add_argument("--config", "-c", required=True, help="Path to config JSON")
    show_parser.add_argument("--category", default="sorumlu", help="Owner tag category (default: sorumlu)")
    show_parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")

    # Template command
    template_parser = subparsers.add_parser("template", help="Create empty Excel template")
    template_parser.add_argument("--output", "-o", default="vm_owners.xlsx", help="Output file")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    if args.command == "template":
        create_template(args.output)
        return

    setup_logging(getattr(args, "verbose", False))

    try:
        vmware_configs = load_config(args.config)
    except Exception as e:
        logger.error(f"Failed to load config: {e}")
        sys.exit(1)

    export_config = ExportConfig(
        owner_tag_category=args.category,
        include_all_tags=getattr(args, "all_tags", False),
    )

    exporter = VMwareTagExporter(vmware_configs, export_config)
    data = exporter.collect_data()

    if args.command == "export":
        exporter.export_to_excel(data, args.output)
        exporter.show_summary(data)
        print(f"✓ Exported to {args.output}")
    elif args.command == "show":
        exporter.show_summary(data)


if __name__ == "__main__":
    main()
